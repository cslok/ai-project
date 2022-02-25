from faker import Faker
import random
from openpyxl import Workbook
import re
from datetime import datetime, timedelta
import common.hobby as hobby
import common.country as country
import time

# EXCEL SETUP====================================================================================================
wb = Workbook()
ws = wb.active
ws.title = "HKG"
ws.append(
    ['BOT_ID', 'AREA', 'Birthday', 'AGE', 'height', 'weight', 'HOBBY', 'interest of Hobby 1', 'HOBBY2',
     'interest of Hobby 2', 'HOBBY3', 'interest of hobby3'])
ws1 = wb.create_sheet("PRC")
ws1.title = "PRC"
ws1.append(
    ['BOT_ID', 'AREA', 'Birthday', 'AGE', 'height', 'weight', 'HOBBY', 'interest of Hobby 1', 'HOBBY2',
     'interest of Hobby 2', 'HOBBY3', 'interest of hobby3'])
ws2 = wb.create_sheet("ROC")
ws2.title = "ROC"
ws2.append(
    ['BOT_ID', 'AREA', 'Birthday', 'AGE', 'height', 'weight', 'HOBBY', 'interest of Hobby 1', 'HOBBY2',
     'interest of Hobby 2', 'HOBBY3', 'interest of hobby3'])
ws3 = wb.create_sheet("KOR")
ws3.title = "KOR"
ws3.append(
    ['BOT_ID', 'AREA', 'Birthday', 'AGE', 'height', 'weight', 'HOBBY', 'interest of Hobby 1', 'HOBBY2',
     'interest of Hobby 2', 'HOBBY3', 'interest of hobby3'])
ws4 = wb.create_sheet("JPN")
ws4.title = "JPN"
ws4.append(
    ['BOT_ID', 'AREA', 'Birthday', 'AGE', 'height', 'weight', 'HOBBY', 'interest of Hobby 1', 'HOBBY2',
     'interest of Hobby 2', 'HOBBY3', 'interest of hobby3'])
ws5 = wb.create_sheet("USA")
ws5.title = "USA"
ws5.append(
    ['BOT_ID', 'AREA', 'Birthday', 'AGE', 'height', 'weight', 'HOBBY', 'interest of Hobby 1', 'HOBBY2',
     'interest of Hobby 2', 'HOBBY3', 'interest of hobby3'])
ws6 = wb.create_sheet("ENG")
ws6.title = "ENG"
ws6.append(
    ['BOT_ID', 'AREA', 'Birthday', 'AGE', 'height', 'weight', 'HOBBY', 'interest of Hobby 1', 'HOBBY2',
     'interest of Hobby 2', 'HOBBY3', 'interest of hobby3'])
ws7 = wb.create_sheet("AUS")
ws7.title = "AUS"
ws7.append(
    ['BOT_ID', 'AREA', 'Birthday', 'AGE', 'height', 'weight', 'HOBBY', 'interest of Hobby 1', 'HOBBY2',
     'interest of Hobby 2', 'HOBBY3', 'interest of hobby3'])
ws8 = wb.create_sheet("RUS")
ws8.title = "RUS"
ws8.append(
    ['BOT_ID', 'AREA', 'Birthday', 'AGE', 'height', 'weight', 'HOBBY', 'interest of Hobby 1', 'HOBBY2',
     'interest of Hobby 2', 'HOBBY3', 'interest of hobby3'])
start = time.process_time()
# COUNTY Setting========================================================================================================

manual_country = ['HKG', 'PRC']
auto_nation = random.choice(list(country.country.keys()))
auto_country = random.choice(list(country.country.values()))
nation = auto_country
hobby_number = random.randint(1, 3)

# population setting
PRC_population = 0
HKG_population = 0
ROC_population = 0
KOR_population = 0
JPN_population = 0
USA_population = 0
ENG_population = 0
AUS_population = 0
population = 0
nation_population_number = country.country_population.get(auto_nation)


# PRC PART==============================================================================================================

class PRC_ID_number(str):

    def __init__(self, id_number):
        super(PRC_ID_number, self).__init__()
        self.id = id_number
        self.area_id = int(self.id[0:6])
        self.birth_year = int(self.id[6:10])
        self.birth_month = int(self.id[10:12])
        self.birth_day = int(self.id[12:14])

    def get_area_name(self):
        """根据区域编号取出区域名称"""
        return country.PRC_AREA_INFO[self.area_id]

    def get_birthday(self):
        """通过身份证号获取出生日期"""
        return "{0}-{1}-{2}".format(self.birth_year, self.birth_month, self.birth_day)

    def get_age(self):
        """通过身份证号获取年龄"""
        now = (datetime.now() + timedelta(days=1))
        year, month, day = now.year, now.month, now.day

        if year == self.birth_year:
            return 0
        else:
            if self.birth_month > month or (self.birth_month == month and self.birth_day > day):
                return year - self.birth_year - 1
            else:
                return year - self.birth_year

    def get_sex(self):
        """通过身份证号获取性别， 女生：0，男生：1"""
        return int(self.id[16:17]) % 2

    def get_check_digit(self):
        """通过身份证号获取校验码"""
        check_sum = 0
        for i in range(0, 17):
            check_sum += ((1 << (17 - i)) % 11) * int(self.id[i])
        check_digit = (12 - (check_sum % 11)) % 11
        return check_digit if check_digit < 10 else 'X'

    @classmethod
    def verify_id(cls, id_number):
        """校验身份证是否正确"""
        if re.match(country.PRC_ID_NUMBER_18_REGEX, id_number):
            check_digit = cls(id_number).get_check_digit()
            return str(check_digit) == id_number[-1]
        else:
            return bool(re.match(country.PRC_ID_NUMBER_15_REGEX, id_number))

    @classmethod
    def generate_id(cls, sex=0):
        """随机生成身份证号，sex = 0表示女性，sex = 1表示男性"""

        # 随机生成一个区域码(6位数)
        id_number = str(random.choice(list(country.PRC_AREA_INFO.keys())))
        # 限定出生日期范围(8位数)
        start, end = datetime.strptime("1949-10-01", "%Y-%m-%d"), datetime.strptime("2001-01-01", "%Y-%m-%d")
        birth_days = datetime.strftime(start + timedelta(random.randint(0, (end - start).days + 1)), "%Y%m%d")
        id_number += str(birth_days)
        # 顺序码(2位数)
        id_number += str(random.randint(10, 99))
        # 性别码(1位数)
        id_number += str(random.randrange(sex, 10, step=2))
        # 校验码(1位数)
        return id_number + str(cls(id_number).get_check_digit())


# COMMON=================================================================================================================

class process_bot(eval(auto_nation)):
    def country_popuplation_commend(self):
        Country_population = eval(auto_nation + '_population')
        return Country_population

    def BOT_ID(self):
        BOT_ID = auto_nation + '-0-' + Country_population
        return BOT_ID

    def nation_population_number(self):
        nation_population_number = country.country_population.get(auto_nation)
        return nation_population_number

    def nationOnsheet(self):
        if auto_nation == 'HKG':
            sheetpage = ws.append
            return sheetpage
        elif auto_nation == 'PRC':
            sheetpage = ws1.append
            return sheetpage
        elif auto_nation == 'ROC':
            sheetpage = ws2.append
            return sheetpage
        elif auto_nation == 'KOR':
            sheetpage = ws3.append
            return sheetpage
        elif auto_nation == 'JPN':
            sheetpage = ws4.append
            return sheetpage
        elif auto_nation == 'USA':
            sheetpage = ws5.append
            return sheetpage
        elif auto_nation == 'ENG':
            sheetpage = ws6.append
            return sheetpage
        elif auto_nation == 'AUS':
            sheetpage = ws7.append
            return sheetpage
        elif auto_nation == 'RUS':
            sheetpage = ws8.append
            return sheetpage

    def education_level(self):
        if auto_nation == auto_nation:
            BOT_Education_level = eval('country.' + auto_nation + '_education_level')

            return str(BOT_Education_level)

    def manual_country_bot_gen(self):
        if auto_nation == list[manual_country]:
            if auto_nation == 'PRC':
                mode = 1
                return mode
            else:
                if auto_nation == 'HKG':
                    mode = 2
                    return mode
        else:
            mode = 0
            return mode

    def population_limit(auto_nation):
        if process_bot(auto_nation).country_population == process_bot(auto_nation).nation_population_number:
            if mode == 0:
                del country.country[auto_nation]
                return print(auto_nation, 'work is completed!')
            if mode >= 1:
                return print(auto_nation, 'work is completed!')


class bot_by_area(mode):
    def full_auto_country_gen_bot(self):
        if mode == 0:
            fake = Faker(nation)
            return fake
        pass

    def PRCID(Self):
        if mode == 1:
            PRCIDnumber = PRC_ID_number.generate_id(random.randint(0, 1))
            return PRCIDnumber
        else:
            pass

    def birthday(mode):
        if mode == 1:
            BOT_birthday = PRC_ID_number(PRCIDnumber).get_birthday()
            return BOT_birthday
        else:
            randomGenerator = random.random()
            dateFormat = '%Y/%m/%d'

            startTime = time.mktime(time.strptime("1970/03/01", dateFormat))
            endTime = time.mktime(time.strptime("2001/01/01", dateFormat))

            randomTime = startTime + randomGenerator * (endTime - startTime)
            BOT_birthday = time.strftime(dateFormat, time.localtime(randomTime))
            return BOT_birthday

    def gender(mode):
        if mode == 1:
            if PRC_ID_number(PRCIDnumber).get_sex() == 1:
                gender = 'male'
                return gender
            else:
                gender = 'female'
                return gender
        else:
            gender = random.choice(['male', 'female'])
        return gender

    def bot_area(mode):
        if mode == 1:
            BOT_AREA = PRC_ID_number(PRCIDnumber).get_area_name()
            return BOT_AREA
        elif mode == 2:
            BOT_AREA = random.choice(country.HKG_area_code)
            return BOT_AREA
        else:
            BOT_AREA = fake.city_name(auto_country)
            return BOT_AREA


class hobby(hobby_number):

    def hobby_choice(self):
        if hobby_number == 1:
            interest = 1
            BOT_hobby = random.choice(hobby.normal_hobby)
            interest2 = 0
            BOT_HOBBY2 = "沒有"
            interest3 = 0
            BOT_HOBBY3 = "沒有"
            return [BOT_hobby, interest, BOT_HOBBY2, interest2, BOT_HOBBY3, interest3]
        if hobby_number == 2:
            interest = random.random()
            interest2 = 1 - interest
            BOT_hobby = random.choice(hobby.normal_hobby)
            BOT_HOBBY2 = random.choice(hobby.normal_hobby2)
            interest3 = 0
            BOT_HOBBY3 = "沒有"
            return [BOT_hobby, interest, BOT_HOBBY2, interest2, BOT_HOBBY3, interest3]
        if hobby_number == 3:
            interest = random.random()
            interest2 = 1 - interest
            interest = random.random()
            interest3 = 1 - interest - interest2
            BOT_hobby = random.choice(hobby.normal_hobby)
            BOT_HOBBY2 = random.choice(hobby.outdoor_hobby)
            BOT_HOBBY3 = random.choice(hobby.normal_hobby2)
            return [BOT_hobby, interest, BOT_HOBBY2, interest2, BOT_HOBBY3, interest3]


class heightNweight(auto_nation, gender):
    def height(auto_nation, gender):
        height_L = eval('country.' + auto_nation + "_" + gender + '_height_H')
        height_H = eval('country.' + auto_nation + "_" + gender + '_height_H')
        BOT_height = random.uniform(height_L, height_H)
        return BOT_height

    def weight(auto_nation, gender):
        weight_L = eval('country.' + auto_nation + "_" + gender + '_height_H')
        weight_H = eval('country.' + auto_nation + "_" + gender + '_height_H')
        BOT_weight = random.uniform(weight_L, weight_H)
        return BOT_weight


#bot generator zone




for population in range(country.total_population):
    if manual_country_bot_gen.mode == 1:
        for country_population in range(nation_population_number):
                print('mode 1 check')
                # gen PRC BOT by population model
    elif manual_country_bot_gen.mode == 2:
        for country_population in range(nation_population_number):
            print('mode 2 check')  # gen HKG bot by population model
        else:
            print("faker moce check complete")
                #ID.BOT_ID, address(mode))


process_bot()

# by country random bot=============================================================================================

# bot birthday =====================================================================================================
######hobby setting=================================================================================================


wb.save(filename='bot_detail test.csv')
