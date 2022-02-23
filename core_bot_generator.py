from faker import Faker
import random
from openpyxl import Workbook
import re
from datetime import datetime, timedelta
import common.hobby as hobby
import common.country as country
import time

# EXCEL SETUP====================================================================================================
wb = Workbook()
ws = wb.active
wb.create_sheet("HKG")
ws.title = "HKG"
ws.append(
    ['Date', 'fee_location_type', 'item', 'project_ID', 'currency', 'Price', 'Amount', 'HOBBY', 'HOBBY2', 'SPORT'])
ws1 = wb.create_sheet("PRC")
ws1.title = "PRC"
ws1.append(
    ['Date', 'fee_location_type', 'item', 'project_ID', 'currency', 'Price', 'Amount', 'HOBBY', 'HOBBY2', 'SPORT'])
ws2 = wb.create_sheet("ROC")
ws2.title = "ROC"
ws2.append(
    ['Date', 'fee_location_type', 'item', 'project_ID', 'currency', 'Price', 'Amount', 'HOBBY', 'HOBBY2', 'SPORT'])
ws3 = wb.create_sheet("KOR")
ws3.title = "KOR"
ws3.append(
    ['Date', 'fee_location_type', 'item', 'project_ID', 'currency', 'Price', 'Amount', 'HOBBY', 'HOBBY2', 'SPORT'])
ws4 = wb.create_sheet("JPN")
ws4.title = "JPN"
ws4.append(
    ['Date', 'fee_location_type', 'item', 'project_ID', 'currency', 'Price', 'Amount', 'HOBBY', 'HOBBY2', 'SPORT'])
ws5 = wb.create_sheet("USA")
ws5.title = "USA"
ws5.append(
    ['Date', 'fee_location_type', 'item', 'project_ID', 'currency', 'Price', 'Amount', 'HOBBY', 'HOBBY2', 'SPORT'])
ws6 = wb.create_sheet("ENG")
ws6.title = "ENG"
ws6.append(
    ['Date', 'fee_location_type', 'item', 'project_ID', 'currency', 'Price', 'Amount', 'HOBBY', 'HOBBY2', 'SPORT'])
ws7 = wb.create_sheet("AUS")
ws7.title = "AUS"
ws7.append(
    ['Date', 'fee_location_type', 'item', 'project_ID', 'currency', 'Price', 'Amount', 'HOBBY', 'HOBBY2', 'SPORT'])
start = time.process_time()
# COUNTY Setting========================================================================================================
manual_country = ['HKG', 'PRC']

# population setting
PRC_population = 0
HKG_population = 0
ROC_population = 0
KOR_population = 0
JPN_population = 0
USA_population = 0
ENG_population = 0
AUS_population = 0
population = 0


# PRC PART==============================================================================================================

class PRC_ID_number(str):

    def __init__(self, id_number):
        super(PRC_ID_number, self).__init__()
        self.id = id_number
        self.area_id = int(self.id[0:6])
        self.birth_year = int(self.id[6:10])
        self.birth_month = int(self.id[10:12])
        self.birth_day = int(self.id[12:14])

    def get_area_name(self):
        """根据区域编号取出区域名称"""
        return country.PRC_AREA_INFO[self.area_id]

    def get_birthday(self):
        """通过身份证号获取出生日期"""
        return "{0}-{1}-{2}".format(self.birth_year, self.birth_month, self.birth_day)

    def get_age(self):
        """通过身份证号获取年龄"""
        now = (datetime.now() + timedelta(days=1))
        year, month, day = now.year, now.month, now.day

        if year == self.birth_year:
            return 0
        else:
            if self.birth_month > month or (self.birth_month == month and self.birth_day > day):
                return year - self.birth_year - 1
            else:
                return year - self.birth_year

    def get_sex(self):
        """通过身份证号获取性别， 女生：0，男生：1"""
        return int(self.id[16:17]) % 2

    def get_check_digit(self):
        """通过身份证号获取校验码"""
        check_sum = 0
        for i in range(0, 17):
            check_sum += ((1 << (17 - i)) % 11) * int(self.id[i])
        check_digit = (12 - (check_sum % 11)) % 11
        return check_digit if check_digit < 10 else 'X'

    @classmethod
    def verify_id(cls, id_number):
        """校验身份证是否正确"""
        if re.match(country.PRC_ID_NUMBER_18_REGEX, id_number):
            check_digit = cls(id_number).get_check_digit()
            return str(check_digit) == id_number[-1]
        else:
            return bool(re.match(country.PRC_ID_NUMBER_15_REGEX, id_number))

    @classmethod
    def generate_id(cls, sex=0):
        """随机生成身份证号，sex = 0表示女性，sex = 1表示男性"""

        # 随机生成一个区域码(6位数)
        id_number = str(random.choice(list(country.PRC_AREA_INFO.keys())))
        # 限定出生日期范围(8位数)
        start, end = datetime.strptime("1949-10-01", "%Y-%m-%d"), datetime.strptime("2001-01-01", "%Y-%m-%d")
        birth_days = datetime.strftime(start + timedelta(random.randint(0, (end - start).days + 1)), "%Y%m%d")
        id_number += str(birth_days)
        # 顺序码(2位数)
        id_number += str(random.randint(10, 99))
        # 性别码(1位数)
        id_number += str(random.randrange(sex, 10, step=2))
        # 校验码(1位数)
        return id_number + str(cls(id_number).get_check_digit())


# COMMON PRE-CONFIG=====================================================================================================
# HOBBY=================================================================================================================
BOT_hobby = random.choice(hobby.normal_hobby)
BOT_hobby2 = random.choice(hobby.outdoor_hobby)
BOT_HOBBY3 = random.choice(hobby.normal_hobby2)


# COMMON================================================================================================================
def getRandomDate(startDate, endDate):
    randomGenerator = random.random()
    dateFormat = '%Y/%m/%d'

    startTime = time.mktime(time.strptime(startDate, dateFormat))
    endTime = time.mktime(time.strptime(endDate, dateFormat))

    randomTime = startTime + randomGenerator * (endTime - startTime)
    randomDate = time.strftime(dateFormat, time.localtime(randomTime))
    return randomDate


# bot generator zone====================================================================================================

while population < country.total_population:

    # by country random bot=============================================================================================
    auto_country = random.choice(list(country.country.values()))
    auto_nation = random.choice(list(country.country.keys()))
    nation = auto_country
    # bot birthday =====================================================================================================
    BOT_birthday = getRandomDate("1970/03/01", "2001/01/01")
    ######hobby setting=================================================================================================
    hobby_number = random.randint(1, 3)
    interest = hobby_number
    if interest == 1:
        interest = 1
        BOT_hobby = random.choice(hobby.normal_hobby)
        BOT_HOBBY2 = "沒有"
        BOT_HOBBY3 = "沒有"
    elif interest == 2:
        interest = random.random()
        interest2 = 1 - interest
        BOT_hobby = random.choice(hobby.normal_hobby)
        BOT_HOBBY2 = random.choice(hobby.normal_hobby2)
        BOT_HOBBY3 = "沒有"
    elif interest == 3:
        interest = random.random()
        interest2 = 1 - interest
        interest = random.random()
        interest3 = 1 - interest - interest2
        BOT_hobby = random.choice(hobby.normal_hobby)
        BOT_hobby2 = random.choice(hobby.outdoor_hobby)
        BOT_HOBBY3 = random.choice(hobby.normal_hobby2)

    # Gender setting====================================================================================================
    gender = random.choice("男")

    if nation in manual_country:
        if auto_nation == 'PRC':
            if PRC_population < country.country_population['PRC']:
                PRC_population += 1
                population += 1
                PRCIDnumber = PRC_ID_number.generate_id(random.randint(0, 1))
                PRC_BOT_ID = 'PRC-0-' + str(PRC_population)
                PRC_education_level = random.choice(country.PRC_education_level)

                if PRC_ID_number(PRCIDnumber).get_sex() == 1:
                    BOT_GENDER = '男'
                    BOT_HEIGHT = random.uniform(country.PRC_male_height_H, country.PRC_male_height_L)
                    BOT_WEIGHT = random.uniform(country.PRC_male_weight_H, country.PRC_male_weight_L)
                else:
                    BOT_GENDER = '女'
                    BOT_HEIGHT = random.uniform(country.PRC_female_height_H, country.PRC_female_height_L)
                    BOT_WEIGHT = random.uniform(country.PRC_female_weight_H, country.PRC_female_weight_L)

                ws.append([PRC_BOT_ID, BOT_GENDER, PRC_ID_number(PRCIDnumber).get_birthday(), BOT_WEIGHT, BOT_HEIGHT,
                           PRC_ID_number(PRCIDnumber).get_age(), PRC_ID_number(PRCIDnumber).get_area_name(), BOT_hobby,
                           BOT_hobby2])
                if PRC_population == country.country_population['PRC']-1:
                    print('中國地區已完成製造 , China is complete')
                world_progress = population / country.total_population
                world_progress_percentage = round(world_progress * 100, 4)
                print(PRC_BOT_ID, ':Worldwide population is', population, 'of', country.total_population,
                      'Worldwide is now in', world_progress_percentage, '%',time.process_time() - start)

            else:
                del country.country.values['PRC']
                del country.country.keys['PRC']


#Hong kong==============================================================================================================
        else:
            if auto_nation == 'HKG':
                if HKG_population < country.country_population['HKG']:
                    HKG_BOT_ID = 'HKG-0-' + str(HKG_population)
                    HKG_population += 1
                    population += 1
                    HKG_education_level = random.choice(country.HKG_education_level)
                    if gender == '男':
                        HKG_height = random.uniform(country.HKG_male_height_H, country.HKG_male_height_L)
                        HKG_weight = random.uniform(country.HKG_male_weight_H, country.HKG_male_weight_L)
                    else:
                        HKG_height = random.uniform(country.HKG_female_height_H, country.HKG_female_height_L)
                        HKG_weight = random.uniform(country.HKG_female_weight_H, country.HKG_female_weight_L)

                    HKG_area = random.choice(country.HKG_area_code)
                    ws.append([HKG_BOT_ID, gender, BOT_birthday, HKG_area, BOT_hobby, BOT_hobby2,
                               BOT_HOBBY3])
                    if HKG_population == country.country_population['HKG']-1:
                        print('香港地區已完成製造 , Hong Kong is complete')
                    world_progress = population / country.total_population
                    world_progress_percentage = round(world_progress * 100, 4)
                    print(HKG_BOT_ID, ':Worldwide population is', population, 'of', country.total_population,
                          'Worldwide is now in', world_progress_percentage, '%',time.process_time() - start)
                else:
                    del country.country.values['HKG']
                    del country.country.keys['HKG']

    elif nation not in manual_country:

        # use country.py nation to format page ,count population , gen id/
        fake = Faker(nation)
        if auto_nation == 'ROC':
            if ROC_population < country.country_population['ROC']:
                ROC_BOT_ID = 'ROC-0-' + str(ROC_population)
                ROC_population += 1
                population += 1
                ROC_education_level = random.choice(country.ROC_education_level)
                if gender == '男':
                    ROC_height = random.uniform(country.ROC_male_height_H, country.ROC_male_height_L)
                    ROC_weight = random.uniform(country.ROC_male_weight_H, country.ROC_male_weight_L)
                else:
                    ROC_height = random.uniform(country.ROC_female_height_H, country.ROC_female_height_L)
                    ROC_weight = random.uniform(country.ROC_female_weight_H, country.ROC_female_height_L)
                ws2.append([ROC_BOT_ID, gender, BOT_birthday, fake.address(), BOT_hobby, BOT_hobby2,
                            BOT_HOBBY3])
                if ROC_population == country.country_population['ROC']-1:
                    print('台灣地區已完成製造 , Taiwan is complete')
                world_progress = population / country.total_population
                world_progress_percentage = round(world_progress * 100, 4)
                print(ROC_BOT_ID, ':Worldwide population is', population, 'of', country.total_population,
                      'Worldwide is now in', world_progress_percentage, '%',time.process_time() - start)
            else:
                del country.country.values['zh_TW']
                del country.country.keys['ROC']

    # South Korea ======================================================================================================
        elif auto_nation == 'KOR':
            if auto_nation == 'KOR':
                if KOR_population < country.country_population['KOR']:
                    KOR_BOT_ID = 'KOR-0-' + str(KOR_population)
                    KOR_population += 1
                    population += 1
                    KOR_education_level = random.choice(country.KOR_education_level)
                    if gender == '男':
                        KOR_height = random.uniform(country.KOR_male_height_H, country.KOR_male_height_L)
                        KOR_weight = random.uniform(country.KOR_male_weight_H, country.KOR_male_weight_L)
                    else:
                        KOR_height = random.uniform(country.KOR_female_height_H, country.KOR_female_height_L)
                        KOR_weight = random.uniform(country.KOR_female_weight_H, country.KOR_female_weight_L)
                    ws3.append([KOR_BOT_ID, gender, BOT_birthday, fake.address(), BOT_hobby, BOT_hobby2,
                                BOT_HOBBY3])
                    if KOR_population == country.country_population['KOR']-1:
                        print('南韓地區已完成製造 , South Korea is complete')
                    world_progress = population / country.total_population
                    world_progress_percentage = round(world_progress * 100, 4)
                    print(KOR_BOT_ID, ':Worldwide population is', population, 'of', country.total_population,
                          'Worldwide is now in', world_progress_percentage, '%',time.process_time() - start)

                else:
                    del country.country.values['ko_KR']
                    del country.country.keys['KOR']


        # Japan=================================================================================================================
        elif auto_nation == 'JPN':
            if auto_nation == 'JPN':
                if JPN_population < country.country_population['JPN']:
                    JPN_BOT_ID = 'JPN-0-' + str(JPN_population)
                    JPN_population += 1
                    population += 1
                    JPN_education_level = random.choice(country.JPN_education_level)
                    if gender == '男':
                        JPN_height = random.uniform(country.JPN_male_height_H, country.JPN_male_height_L)
                        JPN_weight = random.uniform(country.JPN_male_weight_H, country.JPN_male_weight_L)
                    else:
                        JPN_height = random.uniform(country.JPN_female_height_H, country.JPN_female_height_L)
                        JPN_weight = random.uniform(country.JPN_female_weight_H, country.JPN_female_weight_L)
                    ws4.append([JPN_BOT_ID, gender, BOT_birthday, fake.address(), BOT_hobby, BOT_hobby2,
                                BOT_HOBBY3])
                    if JPN_population == country.country_population['JPN']-1:
                        print('日本地區已完成製造 , JAPAN is complete')
                    world_progress = population / country.total_population
                    world_progress_percentage = round(world_progress * 100, 4)
                    print(JPN_BOT_ID, ':Worldwide population is', population, 'of', country.total_population,
                          'Worldwide is now in', world_progress_percentage, '%',time.process_time() - start)

                else:
                    del country.country.values['ja_JP']
                    del country.country.keys['JPN']
                    print('日本地區已完成製造 , JAPAN is complete')

#USA====================================================================================================================
        elif auto_nation == 'USA':
            if auto_nation == 'USA':
                if USA_population < country.country_population['USA']:
                    USA_BOT_ID = 'USA-0-' + str(USA_population)
                    USA_population += 1
                    population += 1
                    USA_education_level = random.choice(country.USA_education_level)
                    if gender == '男':
                        USA_height = random.uniform(country.USA_male_height_H, country.USA_male_height_L)
                        USA_weight = random.uniform(country.USA_male_weight_H, country.USA_male_weight_L)
                    else:
                        USA_height = random.uniform(country.USA_female_height_H, country.USA_female_height_L)
                        USA_weight = random.uniform(country.USA_female_weight_H, country.USA_female_height_L)
                    ws5.append([USA_BOT_ID, gender, BOT_birthday, fake.address(), BOT_hobby, BOT_hobby2,
                                BOT_HOBBY3])
                    if USA_population == country.country_population['USA']-1:
                        print('美國地區已完成製造 ,United States of America is complete')
                    world_progress = population / country.total_population
                    world_progress_percentage = round(world_progress * 100, 4)
                    print(USA_BOT_ID, ':Worldwide population is', population, 'of', country.total_population,
                          'Worldwide is now in', world_progress_percentage, '%',time.process_time() - start)

                else:
                    del country.country.values['en_us']
                    del country.country.keys['USA']


#UK====================================================================================================================
        elif auto_nation == 'ENG':
            if auto_nation == 'ENG':
                if ENG_population < country.country_population['ENG']:
                    ENG_BOT_ID = 'ENG-0-' + str(ENG_population)
                    ENG_population += 1
                    population += 1
                    ENG_education_level = random.choice(country.ENG_education_level)
                    if gender == '男':
                        ENG_height = random.uniform(country.ENG_male_height_H, country.ENG_male_height_L)
                        ENG_weight = random.uniform(country.ENG_male_weight_H, country.ENG_male_weight_L)
                    else:
                        ENG_height = random.uniform(country.ENG_female_height_H, country.ENG_female_height_L)
                        ENG_weight = random.uniform(country.ENG_female_weight_H, country.ENG_female_height_L)
                    ws6.append([ENG_BOT_ID, gender, BOT_birthday, fake.address(), BOT_hobby, BOT_hobby2,
                                BOT_HOBBY3])
                    if ENG_population == country.country_population['ENG'] -1:
                        print('英國地區已完成製造 ,United Kingdom of America is complete')
                    world_progress = population / country.total_population
                    world_progress_percentage = round(world_progress * 100, 4)
                    print(ENG_BOT_ID, ':Worldwide population is', population, 'of', country.total_population,
                          'Worldwide is now in', world_progress_percentage, '%',time.process_time() - start)
                else:
                    del country.country.values['en_GB']
                    del country.country.keys['ENG']
#AU=====================================================================================================================
        else:
            if auto_nation == 'AUS':
                if AUS_population < country.country_population['AUS']:
                    AUS_BOT_ID = 'AUS-0-' + str(AUS_population)
                    AUS_population += 1
                    population += 1
                    AUS_education_level = random.choice(country.AUS_education_level)
                    if gender == '男':
                        AUS_height = random.uniform(country.AUS_male_height_H, country.AUS_male_height_L)
                        AUS_weight = random.uniform(country.AUS_male_weight_H, country.AUS_male_weight_L)
                    else:
                        AUS_height = random.uniform(country.AUS_female_height_H, country.AUS_female_height_L)
                        AUS_weight = random.uniform(country.AUS_female_weight_H, country.AUS_female_height_L)
                    ws7.append([AUS_BOT_ID, gender, BOT_birthday, fake.address(), BOT_hobby, BOT_hobby2,
                                BOT_HOBBY3])
                    if AUS_population == country.country_population['AUS']-1:
                        print('澳州地區已完成製造 ,Australia is complete')
                    world_progress = population / country.total_population
                    world_progress_percentage = round(world_progress * 100, 4)
                    print(AUS_BOT_ID, ':Worldwide population is', population, 'of', country.total_population,
                          'Worldwide is now in', world_progress_percentage, '%',time.process_time() - start)
                else:
                    del country.country.values['en_AU']
                    del country.country.keys['AUS']

    else:
        print('mission complete',time.process_time() - start)

wb.save(filename='bot_detail.csv')
